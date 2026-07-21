"""Tabular loaders (CSV, XLSX)."""

from __future__ import annotations

from pathlib import Path

from langchain_core.documents import Document

from app.ingestion.loaders._common import tag_docs


def load_csv(path: Path) -> list[Document]:
    from langchain_community.document_loaders import CSVLoader

    docs = CSVLoader(str(path), encoding="utf-8").load()
    return tag_docs(docs, source_type="CSV", content_type="TEXT")


def load_xlsx(path: Path) -> list[Document]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    docs: list[Document] = []
    for sheet in wb.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            docs.append(
                Document(
                    page_content="\n".join(rows),
                    metadata={"format": "xlsx", "section": sheet.title, "sheet": sheet.title},
                )
            )
    return tag_docs(docs, source_type="XLSX", content_type="TEXT")
